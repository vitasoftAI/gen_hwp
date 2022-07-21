import pandas as pd
import openpyxl
import pickle
import random
from datetime import datetime
from random import randrange
from datetime import timedelta
import argparse
import os
from bs4 import BeautifulSoup
from tqdm import tqdm

class GenHwp:
    def __init__(self, file_path):
        print("\r 변환 준비중 ...")
        # wb = openpyxl.load_workbook(file_path)
        # self.sheet_name_list = wb.sheetnames
        self.d1 = datetime.strptime('1/1/2000', '%m/%d/%Y')
        self.d2 = datetime.strptime('7/15/2022', '%m/%d/%Y')
        self.total_dict, self.excel_dict = self.load_db(file_path)
        print("\r 준비완료 ... 변환 시작 ..")

    def load_db(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet_name_list = wb.sheetnames
        self.sheet_name_list = sheet_name_list
        with open('sheet_name_list.pickle', 'wb') as f:
            f.write(pickle.dumps(sheet_name_list))
        excel_dict = pd.read_excel(file_path,sheet_name=wb.sheetnames, header=None,engine='openpyxl')

        for key, item in excel_dict.items():
            if key == 'VDPSpec' or key == 'VDataSpec':
                continue
            excel_dict[key] = excel_dict[key].dropna()
        data_field_dict = {}
        data_format_dict = {}
        total_dict = {}
        data_field = excel_dict['VDPSpec']
        data_format = excel_dict['VDataSpec']
        # 필드 및 포멧형식 만들기
        for i in range(len(data_field)):
            data_field_dict[data_field.iloc[i][0]] = data_field.iloc[i][1]
        for j in range(len(data_format)):
            data_format_dict[data_format.iloc[j][0]] = data_format.iloc[j][1]
        # 필드 및 포멧형식 재구성
        for key, val in data_field_dict.items():
            key = key.upper()
            if len(val.split('+')) == 1:
                if val in data_format_dict.keys():
                    total_dict[key] = data_format_dict[val]
            # 여러개를 합친거 처리 필요
            else:
                item_lst = []
                for lst in val.split('+'):
                    if lst in data_format_dict.keys():
                        item_lst.append(data_format_dict[lst])
                total_dict[key] = item_lst
        return total_dict, excel_dict

    def random_date(self, start, end):
        delta = end - start
        int_delta = (delta.days)
        random_second = randrange(int_delta)
        return start + timedelta(days=random_second)
    def number_change(self, num_data):
        while '@AA' in num_data:
            targte_value = random.randint(1, 9)
            num_data = num_data.replace('@AA', str(targte_value), 1)

        while '#AA' in num_data:
            targte_value = random.randint(0, 9)
            num_data = num_data.replace('#AA', str(targte_value), 1)
        return num_data

    def change_data(self, value):
        str_data = ''
        indices = self.excel_dict[value]
        if indices.values.flatten().tolist():
            str_data = random.choice(indices.values.flatten().tolist())
        return str_data
    def edit_content(self, data, src, target):
        return data.replace(str(src).encode(), str(target).encode(), 1)

    def change_str(self,key, value):
        with open('sheet_name_list.pickle', 'rb') as f:
            sheet_name_list = pickle.loads(f.read())

        new_value = ""
        if isinstance(value, list):
            sum_data = ''
            for idx, item in enumerate(value):
                if item in sheet_name_list:
                    str_data = self.change_data(item)
                    sum_data = sum_data + str_data
                    num_data = self.number_change(item)
                    sum_data = sum_data + num_data
            new_value = sum_data
            return new_value

        if 'MM/DD/YYYY' in value:
            new_value = self.random_date(self.d1, self.d2).strftime("%m/%d/%Y")
        if value in sheet_name_list:
            if value == 'date1':
                return self.random_date(self.d1, self.d2).strftime(self.change_data(value))
            else:
                if value.lower() == 'telno1':
                    pass
                str_data = self.change_data(value)
                new_value = self.number_change(str_data)
                if 'fax' in value:
                    new_value = self.total_dict[[item for item in self.total_dict.keys() if 'tel' in item][0]][
                                :-1] + str(
                        random.randint(0, 9))
                return new_value
        num_data = self.number_change(self.total_dict[key])
        new_value = num_data

        try:
            return new_value.upper()
        except Exception as e:
            print(e)
            return new_value

    def generate(self, in_f_path, result_path):
        with open(in_f_path, 'rb') as f:
            data = f.read()
        soup = BeautifulSoup(data, 'html.parser')
        total = len(soup.select("RECTANGLE"))
        for idx, item in tqdm(enumerate(soup.select("RECTANGLE"))):
            if item.char is None:
                continue
            char = item.char.text
            if not '#' in char:
                continue
            data = self.edit_content(data, char,
                                self.change_str(char.strip().upper(), self.total_dict[char.strip().upper()]).replace('&',
                                                                                                           '&amp;'))
        path = result_path

        i = 0
        while True:

            replaced_path = path.replace("#", str(i).zfill(4))
            if not os.path.exists(replaced_path):
                break
            i += 1
        with open(replaced_path, 'wb') as f:
            f.write(data)

parser = argparse.ArgumentParser()
parser.add_argument('-ep','--excel-path', type=str, default="VARIAB_2.xlsx", help='입력 엑셀 경로')
parser.add_argument('-hp','--hml-path', type=str, default="Form_NV_0004.hml", help='입력 hml 경로')
parser.add_argument('-op','--output-path', type=str, default="result#.hml", help='#이 숫자로 바뀐 result 결과값')
opt = parser.parse_args()
howmany = input('몇개를 만드실래요')
gen_hwp = GenHwp(opt.excel_path)

for i in tqdm(range(int(howmany))):
    gen_hwp.generate(opt.hml_path, opt.output_path)
